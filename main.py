import random
import sys

import openpyxl


def gen_score_table(location, sheet1, sheet2, sheet3):
    student_count = 100
    wb = openpyxl.Workbook()
    sh = wb.create_sheet(sheet1)
    sh.cell(1, 1, "poi")
    for i in range(1, student_count + 1):
        sh.cell(i + 1, 1, "酒店%d" % i)

    order = 1
    cybb = ["大床房", "双床房", "豪华大床房", "豪华双床房", "行政大床房", "行政套房", "总统套房", "总统套房带泳池",
            "独立别墅"]
    for s in cybb:
        sh.cell(1, order + 1, s)
        for i in range(1, student_count + 1):
            # 随机一个成绩
            sh.cell(i + 1, order + 1, random.randrange(850, 2000))
        order = order + 1

    sh = wb.create_sheet(sheet2)
    # 学生名字
    sh.cell(1, 1, "poi")
    for i in range(1, student_count + 1):
        sh.cell(i + 1, 1, "酒店%d" % i)

    order = 1
    for s in cybb:
        sh.cell(1, order + 1, s)
        for i in range(1, student_count + 1):
            # 随机一个成绩
            sh.cell(i + 1, order + 1, random.randrange(850, 2000))
        order = order + 1

    sh = wb.create_sheet(sheet3)
    cybb = ["酒店名称", "对应员工", "间夜数量"]
    for i in range(1, 4):
        sh.cell(1, i, cybb[i - 1])

    for i in range(1, student_count + 1):
        sh.cell(i + 1, 1, "酒店%d" % i)

    staff_names = ["小明", "小红", "小亮", "小王", "小李"]
    for i in range(1, student_count + 1):
        sh.cell(i + 1, 2, staff_names[random.randrange(1, len(staff_names))])

    for i in range(1, student_count + 1):
        sh.cell(i + 1, 3, random.randrange(0, 31))

    wb.save(location)


def read_table(location, sheet):
    """
    读取 location 位置的一个名为 sheet 的表格
    :param location: 位置
    :param sheet: sheet名字
    :return: void
    """
    wb = openpyxl.load_workbook(location)
    sh = wb[sheet]
    return sh


def find_major_index(sheet_, kemu):
    """
    :param sheet_: 表格sheet
    :param kemu: 科目字符串
    :return: 某个科目在这个表中的列数
    """
    index_ = 1
    while True:
        ce = sheet_.cell(row=1, column=index_)
        if ce.value == kemu:
            return index_
        elif index_ >= 11 or len(ce.value) == 0:
            return -1
        else:
            index_ = index_ + 1


def find_student_index(sheet__, student):
    index_ = 1
    while True:
        ce = sheet__.cell(row=index_, column=1)
        if ce.value == student:
            return index_
        elif len(ce.value) > 0:
            index_ = index_ + 1
        else:
            return -1


def v_lookup(sheet__, target, cols, no_col):
    """
    简单实现 v_lookup
    :param sheet__: 表
    :param target: 需要寻找的目标
    :param cols: 寻找范围，列表 [1,2,3,4]
    :param no_col: 返回第几列
    :return: v_lookup 的结果
    """
    target_row = -1
    for j in cols:
        for i in range(1, 102):
            ce = sheet__.cell(row=i, column=j)
            if ce.value == target:
                target_row = i
                break
        if target_row != -1:
            break
    if target_row == -1:
        return None
    print(target, '位于', target_row, '列')
    target_col = cols[no_col - 1]
    return sheet__.cell(row=target_row, column=target_col).value


def student_score_avg(sheet, student):
    student_index = find_student_index(sheet, student)
    sum_ = 0
    # break continue
    for i in range(2, 11):
        ce = sheet.cell(row=student_index, column=i)
        sum_ = sum_ + ce.value
    return sum_ / 9


def avg_area(sheet_, i1, j1, i2, j2):
    i_min = min(i1, i2)
    i_max = max(i1, i2)
    j_min = min(j1, j2)
    j_max = max(j1, j2)

    cnt = 0
    sum_ = 0
    for j in range(j_min, j_max + 1):
        for i in range(i_min, i_max + 1):
            ce = sheet_.cell(row=i, column=j)
            sum_ = sum_ + ce.value
            cnt = cnt + 1
    print(cnt, sum_, sum_ / cnt)
    return sum_ / cnt


def v_lookup(sheet__, target, cols, no_col):
    """
    简单实现 v_lookup
    :param sheet__: 表
    :param target: 需要寻找的目标
    :param cols: 寻找范围，列表 [1,2,3,4]
    :param no_col: 返回第几列
    :return: v_lookup 的结果
    """
    target_row = -1
    for j in cols:
        i = 1
        while True:
            ce = sheet__.cell(row=i, column=j)
            if ce.value == target:
                target_row = i
                break
            elif ce.value is not None:
                i = i + 1
            else:
                break
        if target_row != -1:
            break

    if target_row == -1:
        return None
    print(target, '位于', target_row, '行')
    target_col = cols[no_col - 1]
    return sheet__.cell(row=target_row, column=target_col).value


def sum_if(sheet, target_range_col, value_range_col, target):
    sum_ = 0
    r = 1
    while True:
        ce = sheet.cell(column=target_range_col, row=r)
        if ce.value is None:
            break
        else:
            if ce.value == target:
                value_cell = sheet.cell(column=value_range_col, row=r)
                sum_ = sum_ + value_cell.value
            r = r + 1
    return sum_


def chbb(sheet, avalue, bvalue, aim):
    i = find_major_index(sheet, avalue)
    j = 2
    k = find_major_index(sheet, bvalue)
    while True:
        ce = sheet.cell(column=i, row=j)
        if ce.value is not None:
            if ce.value > aim:
                cf = sheet.cell(column=k, row=j)
                print(cf.value)
            j += 1
        else:
            break


def bhnc(sheet, avalue, bvalue):
    i = find_major_index(sheet, avalue)
    j = find_major_index(sheet, bvalue)
    k = 2
    b = 0
    while True:
        ce1 = sheet.cell(column=i, row=k)
        ce2 = sheet.cell(column=j, row=k)
        if ce1.value is not None and ce2.value is not None:
            if ce1.value > ce2.value:
                b += 1
        else:
            break
        k += 1
    return b


def kkbeat(ota1, ota2, goods):
    i = find_major_index(ota1, goods)
    j = 2
    m = find_major_index(ota2, goods)
    bts = 0
    while True:
        ce1 = ota1.cell(column=i, row=j)
        n = find_student_index(ota2, ota1.cell(column=1, row=j).value)
        ce2 = ota2.cell(column=m, row=n)
        if ce1.value is not None and ce2.value is not None:
            if ce1.value > ce2.value:
                bts += 1
        else:
            break
        j += 1
    hh = bts / (j - 1) * 100
    return str(hh) + '%'


def kkbeat2(ota1, ota2, goods):
    ota1_col = find_major_index(ota1, goods)
    ota2_col = find_major_index(ota2, goods)
    ota1_row = 2
    bts = 0
    while True:
        poi_name = ota1.cell(column=1, row=ota1_row).value
        ota2_row = find_student_index(ota2, poi_name)

        ce1 = ota1.cell(sys.maxsize)
        ce2 = ota2.cell(column=ota2_col, row=ota2_row)

        if ce1.value is not None and ce2.value is not None:
            if ce1.value > ce2.value:
                bts += 1
        else:
            break
        ota1_row += 1
    return bts / (ota1_row - 1) * 100


def find_min_demo(l):
    min_ = 999999999
    for i in l:
        if min_ > i:
            min_ = i

    return min_


# 实现一个函数，给定指定的房型，请返回ota1上，该房型最便宜的酒店(poi)名称:
def find_min(ota1, ota2, goods_name):
    goods_name1_col = find_major_index(ota1, goods_name)
    goods_name2_col = find_major_index(ota2, goods_name)
    min1 = sys.maxsize
    min2 = sys.maxsize
    target_row1 = -1
    target_row2 = -1
    for i in range(2, 102):
        ce1 = ota1.cell(column=goods_name1_col, row=i)
        ce2 = ota2.cell(column=goods_name2_col, row=i)
        if min1 > ce1.value:
            min1 = ce1.value
            target_row1 = i
        if min2 > ce2.value:
            min2 = ce2.value
            target_row2 = i
    print(min1, min2)

    if min1 > min2:
        return ota2.cell(column=1, row=target_row2).value
    else:
        return ota1.cell(column=1, row=target_row1).value


# 实现一个函数，计算出指定的某个员工的达标酒店的比例。
# 比例 = 这个员工下达到间夜数量为night的酒店的数量 / 属于这个员工的酒店数量

def compliance_rate(sales_sheet, staff_name, night):
    staff_name_col = find_major_index(sales_sheet, '对应员工')
    night_col = find_major_index(sales_sheet, '间夜数量')
    t = 0
    s = 0
    i = 2
    while True:
        nightce = sales_sheet.cell(column=night_col, row=i)
        staffnamece = sales_sheet.cell(column=staff_name_col, row=i)
        if nightce.value is not None and staffnamece.value == staff_name:
            if nightce.value >= night:
                s += 1
            else:
                t += 1
            i += 1
        elif nightce.value is None or staffnamece.value is None:
            break
        else:
            i += 1

    return s/(s+t)

def compliance_rate2(sales_sheet, staff_name, night):
    staff_name_col = find_major_index(sales_sheet, '对应员工')
    night_col = find_major_index(sales_sheet, '间夜数量')
    all_staff_cnt = 0
    compliance_cnt = 0
    i = 2
    while True:
        night_ce = sales_sheet.cell(column=night_col, row=i)
        staff_name_ce = sales_sheet.cell(column=staff_name_col, row=i)
        if night_ce.value is not None and staff_name_ce.value is not None:
            if staff_name_ce.value == staff_name:
                if night_ce.value >= night:
                    compliance_cnt += 1
                all_staff_cnt += 1
            i += 1
        else:
            break
    return compliance_cnt / all_staff_cnt

if __name__ == '__main__':
    location = "cybb.xlsx"
    sheet_1 = "ota1"
    sheet_2 = "ota2"
    sheet_3 = "sales"
    # gen_score_table(location, sheet_1, sheet_2, sheet_3)

    ota1 = read_table(location, sheet_3)
    cybb = compliance_rate(ota1, "小李", 12)
    print(cybb)